import openpyxl as xl
from openpyxl.chart import PieChart, Reference

def process_workbook(filename):
    work= xl.load_workbook(filename)
    sheet = work["Sheet1"]
    cell = sheet["a1"]
    sheet.cell(1,1)
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        cor_pri = cell.value * 0.9
        cor_pri_cell = sheet.cell(row, 4)
        cor_pri_cell.value = cor_pri
    values = Reference(sheet, min_row= 2,
              max_row= sheet.max_row,
              min_col= 4,
              max_col= 4)
    chart = PieChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'A6')
    work.save(filename)
process_workbook("transactions.xlsx")